"""
Calibración de Gabillon 4.2 (Gibson-Schwartz) — Newton
Parámetros a calibrar: delta0, kappa
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def future_gabillon42(S0, tau, r, delta0, kappa, alpha_hat, sigma1, sigma2, rho):
    B = (1.0 - np.exp(-kappa * tau)) / kappa
    C = (alpha_hat + rho * sigma1 * sigma2 / kappa) * (B - tau) \
      + sigma2**2 * B**2 / (4.0 * kappa)
    ln_F = np.log(S0) - delta0 * B + r * tau + C
    return np.exp(ln_F)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def func_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho):
    delta0 = a[0]
    kappa  = a[1]
    j = 0.0
    for i in range(len(F_mercado)):
        z = F_mercado[i] - future_gabillon42(S0, tau[i], r, delta0, kappa,
                                              alpha_hat, sigma1, sigma2, rho)
        j += z * z
    return j

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def grad_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho):
    n = len(a)
    g = np.zeros(n)
    for j in range(n):
        h = 1e-5 * max(abs(a[j]), 1.0)
        ap = a.copy(); ap[j] += h
        am = a.copy(); am[j] -= h
        g[j] = (func_J(ap, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
              - func_J(am, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)) / (2.0*h)
    return g

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def newton(S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho, a0, niter, eps):

    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Newton, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)

        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
            gm = grad_J(am, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        Z = np.linalg.solve(JG, gk)
        a_new = a - Z

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        Jk = func_J(a_new, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)

        print(' Newton, k: ', k, '  err: ', err, '  J: %.4f' % Jk, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def cargar_datos(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = wb['Spots']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    fecha_obs = rows[0][0]
    S0 = rows[0][1]

    ws2 = wb['Futures']
    rows2 = list(ws2.iter_rows(values_only=True))

    tau_list = []
    F_list = []
    for row in rows2[2:]:
        fecha_venc = row[0]
        precio = row[1]
        if fecha_venc is not None and precio is not None:
            if hasattr(fecha_venc, 'year'):
                days = (fecha_venc - fecha_obs).days
                if days > 0:
                    tau_list.append(days / 365.0)
                    F_list.append(precio)

    return S0, np.array(tau_list), np.array(F_list), fecha_obs

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

if __name__ == '__main__':

    print()
    print(' Calibración Gabillon 4.2 — Newton')
    print(' Parámetros: delta0, kappa')
    print()

    filepath = 'TFM_Assets.xlsx'
    S0, tau, F_mercado, fecha_obs = cargar_datos(filepath)

    # Parámetros fijados
    r = 0.04
    sigma1 = 0.29
    sigma2 = 0.15
    rho = -0.30
    alpha_hat = 0.06

    print(' Fecha: ', fecha_obs.strftime('%Y-%m-%d'))
    print(' S0 = ', S0)
    print(' r = ', r)
    print(' sigma1 = ', sigma1)
    print(' sigma2 = ', sigma2)
    print(' rho = ', rho)
    print(' alpha_hat = ', alpha_hat)
    print(' Futuros: ', len(tau))

    a0 = np.array([0.14, 0.30])
    niter = 50
    eps = 1e-10

    a_opt = newton(S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho,
                    a0, niter, eps)

    delta0_opt = a_opt[0]
    kappa_opt = a_opt[1]
    J_opt = func_J(a_opt, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
    RMSE = np.sqrt(J_opt / len(tau))

    print()
    print(' Resultado:')
    print('   delta0* = %.6f  (%.2f%%)' % (delta0_opt, delta0_opt*100))
    print('   kappa*  = %.6f' % kappa_opt)
    print('   J       = %.6f' % J_opt)
    print('   RMSE    = %.4f' % RMSE)

    # Gráfica
    tau_plot = np.linspace(0.01, max(tau)*1.1, 200)
    F_cal = future_gabillon42(S0, tau_plot, r, delta0_opt, kappa_opt,
                               alpha_hat, sigma1, sigma2, rho)
    F_mod = future_gabillon42(S0, tau, r, delta0_opt, kappa_opt,
                               alpha_hat, sigma1, sigma2, rho)

    plt.figure(figsize=(10, 6))
    plt.plot(tau, F_mercado, 'ro', markersize=5, label='Futuros mercado')
    plt.plot(tau, F_mod, 'bx', markersize=7, label='Modelo calibrado')
    plt.plot(tau_plot, F_cal, 'b-', lw=1.5, alpha=0.7,
             label='Gabillon 4.2 (d0=%.3f, k=%.3f)' % (delta0_opt, kappa_opt))
    plt.axhline(S0, color='green', ls='--', lw=1, label='Spot S0 = %.2f' % S0)
    plt.xlabel('Tiempo a vencimiento (años)')
    plt.ylabel('Precio del futuro')
    plt.title('Calibración Gabillon 4.2 (Newton)')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('calibracion_gabillon42_newton.png', dpi=150)
    plt.show()

    print()
