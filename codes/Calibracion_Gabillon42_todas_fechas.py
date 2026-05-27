"""
Calibración de Gabillon 4.2 (Gibson-Schwartz) — Newton
Parámetros a calibrar: delta0, kappa
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def future_gabillon42(S0, tau, r, delta0, kappa, alpha_hat, sigma1, sigma2, rho):
    B = (1.0 - np.exp(-kappa * tau)) / kappa
    C = (alpha_hat + rho * sigma1 * sigma2 / kappa) * (B - tau) \
      + sigma2**2 * B**2 / (4.0 * kappa)
    ln_F = np.log(S0) - delta0 * B + r * tau + C
    return np.exp(ln_F)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def func_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho):
    delta0 = a[0]
    kappa  = a[1]
    j = 0.0
    for i in range(len(F_mercado)):
        z = F_mercado[i] - future_gabillon42(S0, tau[i], r, delta0, kappa,
                                              alpha_hat, sigma1, sigma2, rho)
        j += z * z
    return j

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def grad_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho):
    n = len(a)
    g = np.zeros(n)
    for j in range(n):
        h = 1e-5 * max(abs(a[j]), 1.0)
        ap = a.copy(); ap[j] += h
        am = a.copy(); am[j] -= h
        g[j] = (func_J(ap, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
              - func_J(am, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)) / (2.0*h)
    return g

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def newton(S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho, a0, niter, eps):

    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Newton, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)

        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
            gm = grad_J(am, S0, tau, F_mercado, r, alpha_hat, sigma1, sigma2, rho)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        Z = np.linalg.solve(JG, gk)
        a_new = a - Z

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        print(' Newton, k: ', k, '  err: ', err, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def cargar_todas_fechas(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws_spot = wb['Spots']
    spots_dict = {}
    for row in ws_spot.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            spots_dict[row[0]] = row[1]

    ws_fut = wb['Futures']
    rows = list(ws_fut.iter_rows(values_only=True))

    fechas_obs = []
    for col_idx in range(0, ws_fut.max_column, 4):
        val = rows[0][col_idx]
        if val is not None and hasattr(val, 'year'):
            fechas_obs.append((val, col_idx))

    datos = []
    for fecha_obs, col_base in fechas_obs:
        S0 = spots_dict.get(fecha_obs, None)
        if S0 is None:
            min_diff = 999
            for f, v in spots_dict.items():
                d = abs((f - fecha_obs).days)
                if d < min_diff: min_diff = d; S0 = v

        tau_list, F_list = [], []
        for row in rows[2:]:
            if col_base < len(row) and col_base + 1 < len(row):
                fv, precio = row[col_base], row[col_base + 1]
                if fv is not None and precio is not None and hasattr(fv, 'year'):
                    days = (fv - fecha_obs).days
                    if days > 0:
                        tau_list.append(days / 365.0)
                        F_list.append(precio)

        if len(tau_list) > 0:
            datos.append({'fecha': fecha_obs, 'S0': S0,
                          'tau': np.array(tau_list), 'F_mercado': np.array(F_list)})

    return datos

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

if __name__ == '__main__':

    print()
    print(' Calibración Gabillon 4.2 — todas las fechas')
    print(' Parámetros: delta0, kappa')
    print()

    filepath = 'TFM_Assets.xlsx'
    datos = cargar_todas_fechas(filepath)

    # Parámetros fijados
    r = 0.04
    sigma1 = 0.29
    sigma2 = 0.15
    rho = -0.30
    alpha_hat = 0.06

    print(' r = ', r)
    print(' sigma1 = ', sigma1)
    print(' sigma2 = ', sigma2)
    print(' rho = ', rho)
    print(' alpha_hat = ', alpha_hat)

    a0 = np.array([0.14, 0.30])    # semilla
    niter = 50
    eps = 1e-10

    resultados = []

    for d in datos:

        print()
        print(' --- %s  S0 = %.2f  (%d futuros) ---'
              % (d['fecha'].strftime('%Y-%m-%d'), d['S0'], len(d['tau'])))

        a_opt = newton(d['S0'], d['tau'], d['F_mercado'], r,
                        alpha_hat, sigma1, sigma2, rho, a0, niter, eps)

        J_opt = func_J(a_opt, d['S0'], d['tau'], d['F_mercado'], r,
                         alpha_hat, sigma1, sigma2, rho)
        RMSE = np.sqrt(J_opt / len(d['tau']))

        resultados.append({
            'fecha': d['fecha'], 'S0': d['S0'],
            'd0': a_opt[0], 'k': a_opt[1], 'RMSE': RMSE,
            'tau': d['tau'], 'F_mercado': d['F_mercado']
        })

        print('   d0* = %.6f  k* = %.6f   RMSE = %.4f' % (a_opt[0], a_opt[1], RMSE))

    # -----------------------------------------------------------------
    # Tabla resumen
    # -----------------------------------------------------------------

    print()
    print(' ═════════════════════════════════════════════════════════════')
    print(' RESUMEN')
    print(' ═════════════════════════════════════════════════════════════')
    print()
    print(' %12s  %7s  %10s  %10s  %8s' % ('Fecha', 'S0', 'δ₀ (%)', 'κ', 'RMSE'))
    print(' %12s  %7s  %10s  %10s  %8s' % ('─'*12, '─'*7, '─'*10, '─'*10, '─'*8))
    for res in resultados:
        print(' %12s  %7.2f  %10.4f  %10.4f  %8.4f'
              % (res['fecha'].strftime('%Y-%m-%d'), res['S0'],
                 res['d0']*100, res['k'], res['RMSE']))
    print()
    print(' RMSE medio: %.4f' % np.mean([r['RMSE'] for r in resultados]))

    # -----------------------------------------------------------------
    # Gráfica 1: curvas calibradas
    # -----------------------------------------------------------------

    plt.figure(figsize=(12, 7))
    colores = plt.cm.viridis(np.linspace(0, 1, len(resultados)))

    for i, res in enumerate(resultados):
        tau_plot = np.linspace(0.01, max(res['tau'])*1.05, 200)
        F_cal = future_gabillon42(res['S0'], tau_plot, r, res['d0'], res['k'],
                                   alpha_hat, sigma1, sigma2, rho)
        plt.plot(res['tau'], res['F_mercado'], 'o', color=colores[i], markersize=3, alpha=0.5)
        plt.plot(tau_plot, F_cal, '-', color=colores[i], lw=1.5,
                 label='%s (d0=%.3f, k=%.3f)' % (res['fecha'].strftime('%d/%m/%y'), res['d0'], res['k']))

    plt.xlabel('Tiempo a vencimiento (años)')
    plt.ylabel('Precio del futuro')
    plt.title('Gabillon 4.2 — todas las fechas')
    plt.legend(fontsize=8)
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('gabillon42_todas_fechas.png', dpi=150)
    plt.show()

    # -----------------------------------------------------------------
    # Gráfica 2: evolución de parámetros
    # -----------------------------------------------------------------

    fechas = [r['fecha'] for r in resultados]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

    ax1.plot(fechas, [r['d0']*100 for r in resultados], 'bo-', markersize=8)
    ax1.set_ylabel('δ₀ (%)')
    ax1.set_title('Convenience yield inicial')
    ax1.grid(True)
    ax1.tick_params(axis='x', rotation=45)

    ax2.plot(fechas, [r['k'] for r in resultados], 'rs-', markersize=8)
    ax2.set_ylabel('κ')
    ax2.set_title('Velocidad de reversión')
    ax2.grid(True)
    ax2.tick_params(axis='x', rotation=45)

    plt.suptitle('Gabillon 4.2 — evolución de parámetros')
    plt.tight_layout()
    plt.savefig('gabillon42_evolucion_params.png', dpi=150)
    plt.show()

    print()
