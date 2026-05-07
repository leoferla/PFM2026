"""
Calibración de Gabillon 4.3 (Spot + Largo Plazo) — Newton
Parámetros a calibrar: L0, beta
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def futures_price_gabillon43(S, L, tau, beta, sigma_S, sigma_L, rho):
    B = np.exp(-beta * tau)
    v = sigma_S**2 + sigma_L**2 - 2.0 * rho * sigma_S * sigma_L
    A = np.exp(v / (4.0 * beta) * (np.exp(-beta * tau) - np.exp(-2.0 * beta * tau)))
    return A * (S ** B) * (L ** (1.0 - B))

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def func_J(a, S0, tau, F_mercado, sigma_S, sigma_L, rho):
    L0   = a[0]
    beta = a[1]
    j = 0.0
    for i in range(len(F_mercado)):
        z = F_mercado[i] - futures_price_gabillon43(S0, L0, tau[i], beta,
                                                     sigma_S, sigma_L, rho)
        j += z * z
    return j

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def grad_J(a, S0, tau, F_mercado, sigma_S, sigma_L, rho):
    n = len(a)
    g = np.zeros(n)
    for j in range(n):
        h = 1e-5 * max(abs(a[j]), 1.0)
        ap = a.copy(); ap[j] += h
        am = a.copy(); am[j] -= h
        g[j] = (func_J(ap, S0, tau, F_mercado, sigma_S, sigma_L, rho)
              - func_J(am, S0, tau, F_mercado, sigma_S, sigma_L, rho)) / (2.0*h)
    return g

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def newton(S0, tau, F_mercado, sigma_S, sigma_L, rho, a0, niter, eps):

    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Newton, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, tau, F_mercado, sigma_S, sigma_L, rho)

        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, tau, F_mercado, sigma_S, sigma_L, rho)
            gm = grad_J(am, S0, tau, F_mercado, sigma_S, sigma_L, rho)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        Z = np.linalg.solve(JG, gk)
        a_new = a - Z

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        Jk = func_J(a_new, S0, tau, F_mercado, sigma_S, sigma_L, rho)

        print(' Newton, k: ', k, '  err: ', err, '  J: %.4f' % Jk, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def cargar_datos(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = wb['Spots']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    fecha_obs = rows[0][0]
    S0 = rows[0][1]

    ws2 = wb['Futures']
    rows2 = list(ws2.iter_rows(values_only=True))

    tau_list = []
    F_list = []
    for row in rows2[2:]:
        fecha_venc = row[0]
        precio = row[1]
        if fecha_venc is not None and precio is not None:
            if hasattr(fecha_venc, 'year'):
                days = (fecha_venc - fecha_obs).days
                if days > 0:
                    tau_list.append(days / 365.0)
                    F_list.append(precio)

    return S0, np.array(tau_list), np.array(F_list), fecha_obs

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

if __name__ == '__main__':

    print()
    print(' Calibración Gabillon 4.3 — Newton')
    print(' Parámetros: L0, beta')
    print()

    filepath = 'TFM_Assets.xlsx'
    S0, tau, F_mercado, fecha_obs = cargar_datos(filepath)

    # Parámetros fijados
    sigma_S = 0.29
    sigma_L = 0.10
    rho = 0.70

    print(' Fecha: ', fecha_obs.strftime('%Y-%m-%d'))
    print(' S0 = ', S0)
    print(' sigma_S = ', sigma_S)
    print(' sigma_L = ', sigma_L)
    print(' rho = ', rho)
    print(' Futuros: ', len(tau))

    a0 = np.array([66.0, 1.5])
    niter = 50
    eps = 1e-10

    a_opt = newton(S0, tau, F_mercado, sigma_S, sigma_L, rho,
                    a0, niter, eps)

    L0_opt   = a_opt[0]
    beta_opt = a_opt[1]
    J_opt = func_J(a_opt, S0, tau, F_mercado, sigma_S, sigma_L, rho)
    RMSE = np.sqrt(J_opt / len(tau))

    print()
    print(' Resultado:')
    print('   L0*   = %.4f' % L0_opt)
    print('   beta* = %.6f' % beta_opt)
    print('   J     = %.6f' % J_opt)
    print('   RMSE  = %.4f' % RMSE)

    # Gráfica
    tau_plot = np.linspace(0.01, max(tau)*1.1, 200)
    F_cal = futures_price_gabillon43(S0, L0_opt, tau_plot, beta_opt,
                                      sigma_S, sigma_L, rho)
    F_mod = futures_price_gabillon43(S0, L0_opt, tau, beta_opt,
                                      sigma_S, sigma_L, rho)

    plt.figure(figsize=(10, 6))
    plt.plot(tau, F_mercado, 'ro', markersize=5, label='Futuros mercado')
    plt.plot(tau, F_mod, 'bx', markersize=7, label='Modelo calibrado')
    plt.plot(tau_plot, F_cal, 'b-', lw=1.5, alpha=0.7,
             label='Gabillon 4.3 (L0=%.2f, beta=%.3f)' % (L0_opt, beta_opt))
    plt.axhline(S0, color='green', ls='--', lw=1, label='Spot S0 = %.2f' % S0)
    plt.axhline(L0_opt, color='orange', ls='--', lw=1, label='L0* = %.2f' % L0_opt)
    plt.xlabel('Tiempo a vencimiento (años)')
    plt.ylabel('Precio del futuro')
    plt.title('Calibración Gabillon 4.3 (Newton)')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('calibracion_gabillon43_newton.png', dpi=150)
    plt.show()

    print()
