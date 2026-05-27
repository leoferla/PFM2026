"""
Calibración de Gabillon 4.1 — para todas las fechas
F(tau) = S0 * exp((r + c) * tau),  c = Cc - Cy
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def future_gabillon41(S0, tau, r, c):
    return S0 * np.exp((r + c) * tau)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def func_J(a, S0, tau, F_mercado, r):
    c = a[0]
    j = 0.0
    for i in range(len(F_mercado)):
        z = F_mercado[i] - future_gabillon41(S0, tau[i], r, c)
        j += z * z
    return j

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def grad_J(a, S0, tau, F_mercado, r):
    n = len(a)
    g = np.zeros(n)
    for j in range(n):
        h = 1e-5 * max(abs(a[j]), 1.0)
        ap = a.copy(); ap[j] += h
        am = a.copy(); am[j] -= h
        g[j] = (func_J(ap, S0, tau, F_mercado, r)
              - func_J(am, S0, tau, F_mercado, r)) / (2.0*h)
    return g

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def newton(S0, tau, F_mercado, r, a0, niter, eps):

    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Newton, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, tau, F_mercado, r)

        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, tau, F_mercado, r)
            gm = grad_J(am, S0, tau, F_mercado, r)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        Z = np.linalg.solve(JG, gk)
        a_new = a - Z

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        print(' Newton, k: ', k, '  err: ', err, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def cargar_todas_fechas(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws_spot = wb['Spots']
    spots_dict = {}
    for row in ws_spot.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            spots_dict[row[0]] = row[1]

    ws_fut = wb['Futures']
    rows = list(ws_fut.iter_rows(values_only=True))

    fechas_obs = []
    for col_idx in range(0, ws_fut.max_column, 4):
        val = rows[0][col_idx]
        if val is not None and hasattr(val, 'year'):
            fechas_obs.append((val, col_idx))

    datos = []
    for fecha_obs, col_base in fechas_obs:
        S0 = spots_dict.get(fecha_obs, None)
        if S0 is None:
            min_diff = 999
            for f, v in spots_dict.items():
                d = abs((f - fecha_obs).days)
                if d < min_diff: min_diff = d; S0 = v

        tau_list, F_list = [], []
        for row in rows[2:]:
            if col_base < len(row) and col_base + 1 < len(row):
                fv, precio = row[col_base], row[col_base + 1]
                if fv is not None and precio is not None and hasattr(fv, 'year'):
                    days = (fv - fecha_obs).days
                    if days > 0:
                        tau_list.append(days / 365.0)
                        F_list.append(precio)

        if len(tau_list) > 0:
            datos.append({'fecha': fecha_obs, 'S0': S0,
                          'tau': np.array(tau_list), 'F_mercado': np.array(F_list)})

    return datos

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

if __name__ == '__main__':

    print()
    print(' Calibración Gabillon 4.1 — todas las fechas')
    print(' F(tau) = S0 * exp((r + c) * tau),  c = Cc - Cy')
    print()

    filepath = 'TFM_Assets.xlsx'
    datos = cargar_todas_fechas(filepath)

    r = 0.04
    a0 = np.array([0.0])
    niter = 50
    eps = 1e-10

    resultados = []

    for d in datos:

        print()
        print(' --- %s  S0 = %.2f  (%d futuros) ---'
              % (d['fecha'].strftime('%Y-%m-%d'), d['S0'], len(d['tau'])))

        a_opt = newton(d['S0'], d['tau'], d['F_mercado'], r, a0, niter, eps)

        J_opt = func_J(a_opt, d['S0'], d['tau'], d['F_mercado'], r)
        RMSE = np.sqrt(J_opt / len(d['tau']))

        resultados.append({
            'fecha': d['fecha'], 'S0': d['S0'],
            'c': a_opt[0], 'RMSE': RMSE,
            'tau': d['tau'], 'F_mercado': d['F_mercado']
        })

        print('   c* = %.6f  (%.2f%%)   RMSE = %.4f' % (a_opt[0], a_opt[0]*100, RMSE))

    # -----------------------------------------------------------------
    # Tabla resumen
    # -----------------------------------------------------------------

    print()
    print(' ═════════════════════════════════════════════════════')
    print(' RESUMEN Gabillon 4.1')
    print(' ═════════════════════════════════════════════════════')
    print()
    print(' %12s  %7s  %10s  %8s' % ('Fecha', 'S0', 'c (%)', 'RMSE'))
    print(' %12s  %7s  %10s  %8s' % ('─'*12, '─'*7, '─'*10, '─'*8))
    for res in resultados:
        print(' %12s  %7.2f  %10.4f  %8.4f'
              % (res['fecha'].strftime('%Y-%m-%d'), res['S0'],
                 res['c']*100, res['RMSE']))
    print()
    print(' RMSE medio: %.4f' % np.mean([r['RMSE'] for r in resultados]))

    # -----------------------------------------------------------------
    # Gráfica 1: curvas calibradas
    # -----------------------------------------------------------------

    plt.figure(figsize=(12, 7))
    colores = plt.cm.viridis(np.linspace(0, 1, len(resultados)))

    for i, res in enumerate(resultados):
        tau_plot = np.linspace(0.01, max(res['tau'])*1.05, 200)
        F_cal = future_gabillon41(res['S0'], tau_plot, r, res['c'])
        plt.plot(res['tau'], res['F_mercado'], 'o', color=colores[i], markersize=3, alpha=0.5)
        plt.plot(tau_plot, F_cal, '-', color=colores[i], lw=1.5,
                 label='%s (c=%.4f)' % (res['fecha'].strftime('%d/%m/%y'), res['c']))

    plt.xlabel('Tiempo a vencimiento (años)')
    plt.ylabel('Precio del futuro')
    plt.title('Gabillon 4.1 — todas las fechas')
    plt.legend(fontsize=8)
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('gabillon41_todas_fechas.png', dpi=150)
    plt.show()

    # -----------------------------------------------------------------
    # Gráfica 2: evolución de c
    # -----------------------------------------------------------------

    fechas = [r['fecha'] for r in resultados]

    plt.figure(figsize=(10, 5))
    plt.plot(fechas, [r['c']*100 for r in resultados], 'ro-', markersize=8)
    plt.xlabel('Fecha')
    plt.ylabel('c = Cc - Cy  (%)')
    plt.title('Gabillon 4.1 — evolución de c')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('gabillon41_evolucion_c.png', dpi=150)
    plt.show()

    print()
