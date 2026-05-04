"""
Calibración de Gabillon 4.1
F(tau) = S0 * exp((r + Cc - Cy) * tau)
Fijamos Cc = 0, calibramos Cy. (Aunque no es realista)

"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def future_price_with_convenience_yield(S, tau, r, Cc, Cy):
    """De Paper_Gabillon41.py"""
    return S * np.exp((r + Cc - Cy) * tau)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Función objetivo
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def func_J(a, S0, r, tau, F_mercado):
    """
    J(a) = sum (F_mercado - F_modelo)^2
    a = [Cy]
    """
    Cy = a[0]
    Cc = 0.0
    j = 0.0
    for i in range(len(F_mercado)):
        z = F_mercado[i] - future_price_with_convenience_yield(S0, tau[i], r, Cc, Cy)
        j += z * z
    return j

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Gradiente por diferencias finitas centradas
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def grad_J(a, S0, r, tau, F_mercado):
    n = len(a)
    g = np.zeros(n)
    for j in range(n):
        h = 1e-5 * max(abs(a[j]), 1.0)
        ap = a.copy(); ap[j] += h
        am = a.copy(); am[j] -= h
        g[j] = (func_J(ap, S0, r, tau, F_mercado) - func_J(am, S0, r, tau, F_mercado)) / (2.0*h)
    return g

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Newton
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def newton(S0, r, tau, F_mercado, a0, niter, eps):

    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Newton, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, r, tau, F_mercado)

        # Jacobiano de G por diferencias finitas centradas
        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, r, tau, F_mercado)
            gm = grad_J(am, S0, r, tau, F_mercado)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        Z = np.linalg.solve(JG, gk)
        a_new = a - Z

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        print(' Newton, k: ', k, '  err: ', err, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Marquardt
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def marquardt(S0, r, tau, F_mercado, a0, niter, eps):

    gamma = 0.20
    a = a0.copy()
    h = 1e-5
    n = len(a)

    print()
    print(' Marquardt, k: ', 0, '  a: ', a)

    for k in range(1, niter):

        gk = grad_J(a, S0, r, tau, F_mercado)

        # Jacobiano de G por diferencias finitas centradas
        JG = np.zeros((n, n))
        for j in range(n):
            hj = h * max(abs(a[j]), 1.0)
            ap = a.copy(); ap[j] += hj
            am = a.copy(); am[j] -= hj
            gp = grad_J(ap, S0, r, tau, F_mercado)
            gm = grad_J(am, S0, r, tau, F_mercado)
            JG[:,j] = (gp - gm) / (2.0 * hj)

        hk = JG + gamma * np.identity(n)
        pk = -np.linalg.solve(hk, gk)
        a_new = a + pk

        err = np.linalg.norm(a_new - a) / np.linalg.norm(a_new)
        print(' Marquardt, k: ', k, '  err: ', err, '  a: ', a_new)

        if err < eps: break
        a = a_new.copy()

    return a_new

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Cargar datos del Excel
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

def cargar_datos(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Spot (primera fila = fecha más reciente)
    ws = wb['Spots']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    fecha_obs = rows[0][0]
    S0 = rows[0][1]

    # Futuros
    ws2 = wb['Futures']
    rows2 = list(ws2.iter_rows(values_only=True))

    tau_list = []
    F_list = []
    for row in rows2[2:]:
        fecha_venc = row[0]
        precio = row[1]
        if fecha_venc is not None and precio is not None:
            if hasattr(fecha_venc, 'year'):
                days = (fecha_venc - fecha_obs).days
                if days > 0:
                    tau_list.append(days / 365.0)
                    F_list.append(precio)

    return S0, np.array(tau_list), np.array(F_list), fecha_obs

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

if __name__ == '__main__':

    print()
    print(' Calibración Gabillon 4.1')
    print(' F(tau) = S0 * exp((r + Cc - Cy) * tau),  Cc = 0')
    print()

    # -----------------------------------------------------------------
    # Datos
    # -----------------------------------------------------------------

    filepath = 'TFM_Assets.xlsx'
    S0, tau, F_mercado, fecha_obs = cargar_datos(filepath)

    r = 0.03

    print(' Fecha: ', fecha_obs.strftime('%Y-%m-%d'))
    print(' S0 = ', S0)
    print(' r  = ', r)
    print(' Futuros: ', len(tau))
    print()

    for i in range(len(tau)):
        print('   tau = %6.4f   F = %8.4f' % (tau[i], F_mercado[i]))

    # -----------------------------------------------------------------
    # Calibración
    # -----------------------------------------------------------------

    a0 = np.array([0.0])
    niter = 50
    eps = 1e-10

    print()
    print(' Método de Newton')
    a_newton = newton(S0, r, tau, F_mercado, a0, niter, eps)
    Cy_newton = a_newton[0]
    J_newton = func_J(a_newton, S0, r, tau, F_mercado)
    RMSE_newton = np.sqrt(J_newton / len(tau))
    print('   Cy* = %.6f  (%.2f%%)' % (Cy_newton, Cy_newton*100))
    print('   J   = %.6f' % J_newton)
    print('   RMSE = %.4f' % RMSE_newton)

    print()
    print(' Método de Levenberg-Marquardt')
    a_marq = marquardt(S0, r, tau, F_mercado, a0, niter, eps)
    Cy_marq = a_marq[0]
    J_marq = func_J(a_marq, S0, r, tau, F_mercado)
    RMSE_marq = np.sqrt(J_marq / len(tau))
    print('   Cy* = %.6f  (%.2f%%)' % (Cy_marq, Cy_marq*100))
    print('   J   = %.6f' % J_marq)
    print('   RMSE = %.4f' % RMSE_marq)

    # -----------------------------------------------------------------
    # Gráfica
    # -----------------------------------------------------------------

    Cy_opt = Cy_newton
    Cc = 0.0

    tau_plot = np.linspace(0.01, max(tau)*1.1, 200)
    F_calibrada = future_price_with_convenience_yield(S0, tau_plot, r, Cc, Cy_opt)
    F_modelo = future_price_with_convenience_yield(S0, tau, r, Cc, Cy_opt)

    plt.figure(figsize=(10, 6))
    plt.plot(tau, F_mercado, 'ro', markersize=5, label='Futuros mercado')
    plt.plot(tau, F_modelo, 'bx', markersize=7, label='Modelo calibrado')
    plt.plot(tau_plot, F_calibrada, 'b-', lw=1.5, alpha=0.7,
             label='Gabillon 4.1 (Cy = %.4f)' % Cy_opt)
    plt.axhline(S0, color='green', ls='--', lw=1, label='Spot S0 = %.2f' % S0)
    plt.xlabel('Tiempo a vencimiento (años)')
    plt.ylabel('Precio del futuro')
    plt.title('Calibración Gabillon 4.1 - %s' % fecha_obs.strftime('%Y-%m-%d'))
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('calibracion_gabillon41.png', dpi=150)
    plt.show()

    print()
